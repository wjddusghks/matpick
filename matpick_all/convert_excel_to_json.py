#!/usr/bin/env python3
"""
맛픽(Matpick) Excel → JSON 변환 스크립트
사용법: python3 convert_excel_to_json.py matpick_template.xlsx output.json

Excel 시트 구조:
- "입력 가이드" → 무시
- "크리에이터 정보" → creators 배열
- 나머지 시트 → restaurants + visits 배열

출력: 앱이 직접 읽는 `matpick-data.json`과 동일한 JSON 구조
권장 출력 경로: client/src/data/matpick-data.json
"""
import sys
import json
import hashlib
import openpyxl

def make_id(prefix, text):
    """텍스트 기반 고유 ID 생성"""
    h = hashlib.md5(text.encode()).hexdigest()[:8]
    return f"{prefix}_{h}"

def safe_str(val):
    """셀 값을 안전한 문자열로 변환"""
    if val is None:
        return ""
    return str(val).strip()

def safe_float(val):
    """셀 값을 float로 변환 (실패 시 0)"""
    if val is None:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0

def convert(excel_path, output_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    creators = []
    restaurants = []
    visits = []
    restaurant_map = {}  # (식당명, 시리즈) → restaurant_id
    
    # ─── 크리에이터 정보 시트 처리 ───
    if "크리에이터 정보" in wb.sheetnames:
        ws = wb["크리에이터 정보"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            series = safe_str(row[0])
            name = safe_str(row[1]) if len(row) > 1 else ""
            channel = safe_str(row[2]) if len(row) > 2 else ""
            youtube_url = safe_str(row[3]) if len(row) > 3 else ""
            subscribers = safe_str(row[4]) if len(row) > 4 else ""
            profile_img = safe_str(row[5]) if len(row) > 5 else ""
            description = safe_str(row[6]) if len(row) > 6 else ""
            
            creator_id = make_id("creator", series)
            creators.append({
                "id": creator_id,
                "name": name,
                "channelName": channel,
                "profileImage": profile_img,
                "subscribers": subscribers,
                "description": description,
                "youtubeUrl": youtube_url,
                "series": series,
            })
    
    # ─── 각 시리즈 시트 처리 ───
    skip_sheets = {"입력 가이드", "크리에이터 정보"}
    
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        
        ws = wb[sheet_name]
        series_name = sheet_name
        
        # 해당 시리즈의 크리에이터 ID 찾기
        creator_id = ""
        for c in creators:
            if c["series"] == series_name:
                creator_id = c["id"]
                break
        
        if not creator_id:
            creator_id = make_id("creator", series_name)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            episode = safe_str(row[0])       # A: 에피소드
            air_date = safe_str(row[1])       # B: 방영일
            video_title = safe_str(row[2])    # C: 영상 제목
            rest_name = safe_str(row[3])      # D: 식당명
            category = safe_str(row[4])       # E: 카테고리
            region = safe_str(row[5])         # F: 지역
            address = safe_str(row[6])        # G: 상세주소
            status = safe_str(row[7])         # H: 상태
            menu = safe_str(row[8])           # I: 대표메뉴
            lat = safe_float(row[9])          # J: 위도
            lng = safe_float(row[10])         # K: 경도
            video_id = safe_str(row[11])      # L: videoId
            video_url = safe_str(row[12])     # M: videoUrl
            domestic = safe_str(row[13])      # N: 국내/해외
            note = safe_str(row[14]) if len(row) > 14 else ""  # O: 비고
            
            # 보류/선정취소/프롤로그는 식당 없이 visit만 생성
            if status in ("보류", "선정취소", "프롤로그") or not rest_name:
                # visit만 생성 (식당 없음)
                visit_id = make_id("v", f"{series_name}_{episode}_{rest_name}")
                visits.append({
                    "id": visit_id,
                    "restaurantId": "",
                    "creatorId": creator_id,
                    "videoId": video_id,
                    "videoUrl": video_url or (f"https://www.youtube.com/watch?v={video_id}" if video_id else ""),
                    "videoTitle": f"[{series_name} {episode}] {video_title}" if video_title else "",
                    "visitDate": air_date,
                    "episode": episode,
                    "rating": "",
                    "comment": note,
                    "thumbnailUrl": f"https://img.youtube.com/vi/{video_id}/hqdefault.jpg" if video_id else "",
                    "series": series_name,
                    "status": status,
                    "isOverseas": domestic == "해외",
                })
                continue
            
            # 식당 생성 (중복 방지)
            rest_key = (rest_name, series_name, region)
            if rest_key not in restaurant_map:
                rest_id = make_id("r", f"{rest_name}_{region}")
                restaurant_map[rest_key] = rest_id
                restaurants.append({
                    "id": rest_id,
                    "name": rest_name,
                    "region": region,
                    "address": address or region,
                    "category": category,
                    "representativeMenu": menu,
                    "lat": lat,
                    "lng": lng,
                    "imageUrl": "",
                    "isOverseas": domestic == "해외",
                })
            
            rest_id = restaurant_map[rest_key]
            
            # visit 생성
            visit_id = make_id("v", f"{series_name}_{episode}_{rest_name}")
            visits.append({
                "id": visit_id,
                "restaurantId": rest_id,
                "creatorId": creator_id,
                "videoId": video_id,
                "videoUrl": video_url or (f"https://www.youtube.com/watch?v={video_id}" if video_id else ""),
                "videoTitle": f"[{series_name} {episode}] {video_title}" if video_title else "",
                "visitDate": air_date,
                "episode": episode,
                "rating": "",
                "comment": note,
                "thumbnailUrl": f"https://img.youtube.com/vi/{video_id}/hqdefault.jpg" if video_id else "",
                "series": series_name,
                "status": status,
                "isOverseas": domestic == "해외",
            })
    
    # ─── 출력 ───
    output = {
        "creators": creators,
        "restaurants": restaurants,
        "visits": visits,
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"✅ JSON 변환 완료: {output_path}")
    print(f"   크리에이터: {len(creators)}개")
    print(f"   식당: {len(restaurants)}개")
    print(f"   방문(visit): {len(visits)}개")
    print(f"   해외 식당: {sum(1 for r in restaurants if r.get('isOverseas'))}개")
    print(f"   보류/선정취소: {sum(1 for v in visits if v.get('status') in ('보류', '선정취소'))}개")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("사용법: python3 convert_excel_to_json.py <excel_path> <output_json_path>")
        print("예시:   python3 convert_excel_to_json.py matpick_template.xlsx client/src/data/matpick-data.json")
        sys.exit(1)
    
    convert(sys.argv[1], sys.argv[2])
